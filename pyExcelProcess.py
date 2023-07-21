import openpyxl
import sys
import os
import time

class Ele_info:
    def __init__(self, idx):
        self.row = 0        #所在行数
        self.count = 0      #数据id
        self.errcode = 0b00000000   #错误码

class Err_List:
    def __init__(self):
        self.ErrEleList = []    #错误数据列表
    
    def add(self, data):
        self.ErrEleList.append(data)

class MyProcess:
    def __init__(self):
        self.err_list = Err_List()  #错误数据列表

    def MQTTAnalyse(self, path):
        global ErrEleIdx    #错误数据序号
        ErrEleIdx = 0
        
        print("MQTTAnalyse")
        print(path)
        workbook = openpyxl.load_workbook(path)
        print(workbook.sheetnames)
        sheet = workbook["转发数据"]
        #print(sheet.dimensions)

        #算出总行数和数据个数
        idx = 3     #从第三行开始
        row = 1
        errcount = 0    #错误个数

        while sheet.cell(row=idx, column=1).value is not None:
            if sheet.cell(row=idx, column=1).value is not None: 
                #判断此行数据是否有错误
                err = self.JudgeEleError(sheet, idx)
                if err != 0b00000000:
                    errcount += 1
                row += 1
            idx += 1
        row += 1
        ele_count = row - 2
        print(f'总行数：{row}, 数据个数：{ele_count}, 错误个数：{errcount}')
        return errcount

    #row_data[]中，0-19分别为：0数据类型号，1数据类型，2来源设备id， 3来源设备名称， 4数据ID，5数据名称
    #6通信方向号， 7通信方向， 8规约类型好， 9规约类型， 10访问地址， 11传输策略号， 12传输策略， 13传输周期(s)， 14传输周期
    #15数据集索引， 16是否记录号， 17是否记录， 18断线存储时间(s)， 19断线存储时间
    def JudgeEleError(self, sheet, row):
        row_data = []
        for i in sheet.iter_rows(min_row=row, max_row=row, min_col=1, max_col=20):
            for cell in i:
                row_data.append(cell.value)
        print(f'第{row}行数据:{row_data}')
        #开始判断
        err = 0b00000000
        if int(row_data[6]) != 0:
            err |= 0b00000001
        if int(row_data[8]) != 90:
            err |= 0b00000010
        if int(row_data[11]) == 0:      #0周期传输
            if int(row_data[0]) == 70 or int(row_data[0]) == 80 or int(row_data[0]) == 100 or int(row_data[0]) == 110:
                err |= 0b00000100
            if int(row_data[13]) == 0:
                err |= 0b00000100
        elif int(row_data[11]) == 1:    #1即时传输
            if (int(row_data[0] != 100 and int(row_data[0]) != 110)) or int(row_data[16]) != 0:
                err |= 0b00010000
        if int(row_data[16]) == 1:      #断线存储
            if int(row_data[18]) == 0 or row_data[18] is None:
                err |= 0b00001000
        print(f'err:{err:b}\n')

        #将错误码和信息存入eleinfo
        if err != 0b00000000:
            global ErrEleIdx
            eleinfo = Ele_info(ErrEleIdx)
            eleinfo.errcode = err
            eleinfo.row = row
            eleinfo.count = row -2
            ErrEleIdx += 1
            self.err_list.add(eleinfo)

        return err

    def IEC104Analyse(self):
        print("IEC104Analyse")

    def modbusTCPAnalyse(self):
        print("modbusTCPAnalyse")